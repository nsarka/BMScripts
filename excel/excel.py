from openpyxl import Workbook
import pandas as pd

def create_excel_file(name):
    # Create initial excel file
    wb = Workbook()
    ws = wb.active

    ws.append(["Hello"])
    wb.save(name)

def write_excel_file(name, command, stats):
    with pd.ExcelWriter(name, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
        sheet = writer.book.create_sheet(title="OMB")
        sheet.cell(column=1, row=sheet.max_row + 2, value=command)
        stats.to_excel(writer, sheet_name=sheet.title, startrow=sheet.max_row)